import csv
import os
import openpyxl

source = "./Annotation Assignment"
target = os.path.join("Assignments_Combined", "final.xlsx")
empty = os.path.join("Assignments_Combined", "empty.xlsx")
final_wb = openpyxl.load_workbook(empty)

for dirName, subdirList, fileList in os.walk(source):
    for fname in fileList:
        path = os.path.join(dirName, fname)
        if (os.path.splitext(fname)[1] == '.xlsx'):
            try:
                print(fname + "Success")
                wb = openpyxl.load_workbook(path)
                names = wb.get_sheet_names()
                for sheet_name in names:
                    sh = wb[sheet_name]
                    target_sh = final_wb.create_sheet(sheet_name)
                    rows = []
                    for row in sh.iter_rows(min_row=1):
                        row_data = []
                        for cell in row:
                            row_data.append(cell.value)
                        # Creating a list of lists, where each list contain a typical row's data
                        with open(sheet_name+".csv", "w") as f:
                            c = csv.writer(f)
                            for r in sh.rows:
                                c.writerow([cell.value for cell in r])
                            target_sh.append(row_data)
            except:
                print(fname + "Failed")
                pass
        else:
            print(fname)

    final_wb.save(empty+"Test.xlsx")